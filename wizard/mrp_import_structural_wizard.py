# -*- coding: utf-8 -*-
"""Import structural orders from Excel/CSV.

Creates BOM + operations + MO from the provided file.
Optionally creates and confirms a Sale Order with lines (PVP + qty).
"""

from odoo import models, fields, _
from odoo.exceptions import UserError

import base64
import io
import csv
import re
import unicodedata
from datetime import datetime, timedelta

try:
    import xlrd  # For .xls files
except Exception:  # pragma: no cover
    xlrd = None
try:
    import openpyxl  # For .xlsx files
except Exception:  # pragma: no cover
    openpyxl = None


class MrpImportStructuralWizard(models.TransientModel):
    _name = 'mrp.import.structural.wizard'
    _description = 'Importador de Estructural'

    file = fields.Binary(string="Archivo Excel/CSV", required=True)
    filename = fields.Char(string="Nombre de archivo")
    create_quotation = fields.Boolean(
        string="Crear cotizacion",
        default=False,
        help="Si esta activo, crea la cotizacion en borrador y agrega sus lineas.",
    )

    HEADER_ALIASES = {
        'producto': ['producto'],
        'producto_codigo': ['referencia', 'codigo', 'codigo_producto', 'product_code', 'default_code', 'ref'],
        'componentes_producto': ['componentes_producto', 'componentesproducto', 'componentes'],
        'por_consumir': ['por consumir', 'por_consumir', 'porconsumir'],
        'operaciones': ['operaciones', 'operacion', 'operacionesproducto'],
        'centro_trabajo': ['centro de trabajo', 'centro_trabajo', 'centrotrabajo'],
        'cantidad_mo': ['cantidad mo', 'cantidad_mo', 'cantidadmo'],
        'origen': ['origen'],
        'pedido_original': ['pedido_original', 'pedido original', 'pedidooriginal'],
        'id_cliente': ['id cliente', 'id_cliente', 'idcliente'],
        'pvp': ['pvp', 'precio', 'precioventa'],
        'fecha_cotizacion': ['fecha cotizacion', 'fecha_cotizacion', 'fechacotizacion', 'fecha'],
        'largo': ['largo', 'longitud'],
        'ancho': ['ancho', 'anchura'],
        'piezas': ['piezas', 'pieza'],
    }
    BASE_REQUIRED_HEADERS = [
        'producto',
        'componentes_producto',
        'por_consumir',
        'operaciones',
        'centro_trabajo',
        'cantidad_mo',
        'origen',
        'pedido_original',
    ]
    QUOTE_REQUIRED_HEADERS = [
        'id_cliente',
        'pvp',
        'fecha_cotizacion',
        'largo',
        'ancho',
        'piezas',
    ]

    def _read_lines(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Debes adjuntar un archivo para importar."))
        if not self.filename:
            raise UserError(_("Debes indicar el nombre del archivo (filename)."))

        content = base64.b64decode(self.file)
        fname = (self.filename or '').lower()

        if fname.endswith('.csv'):
            csvfile = io.StringIO(content.decode('utf-8-sig'))
            reader = csv.DictReader(csvfile)
            return list(reader)

        if fname.endswith('.xlsx'):
            if openpyxl is None:
                raise UserError(_("No se puede leer Excel .xlsx porque falta la libreria Python 'openpyxl' en el servidor."))
            try:
                workbook = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
            except Exception:
                raise UserError(_("No se pudo leer el archivo .xlsx. Verifica que este bien formado."))
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers_row = next(rows_iter)
            except StopIteration:
                return []
            headers = [str(cell).strip() if cell is not None else '' for cell in headers_row]
            lines = []
            for row in rows_iter:
                line = {}
                for colx in range(len(headers)):
                    value = row[colx] if colx < len(row) else None
                    line[headers[colx]] = str(value).strip() if value is not None else ''
                lines.append(line)
            return lines

        if xlrd is None:
            raise UserError(_("No se puede leer Excel .xls porque falta la libreria Python 'xlrd' en el servidor."))
        try:
            workbook = xlrd.open_workbook(file_contents=content)
        except Exception:
            raise UserError(_("No se pudo leer el archivo. Verifica que sea .xls o .csv."))
        sheet = workbook.sheet_by_index(0)
        headers = [str(cell.value).strip() for cell in sheet.row(0)]
        lines = []
        for rowx in range(1, sheet.nrows):
            line = {headers[colx]: str(sheet.cell(rowx, colx).value).strip() for colx in range(len(headers))}
            lines.append(line)
        return lines

    def _normalize_header(self, value):
        text = str(value or '').strip().lower()
        text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
        return re.sub(r'[^a-z0-9]', '', text)

    def _resolve_headers(self, lines):
        if not lines:
            return {}
        raw_headers = list(lines[0].keys())
        header_map = {}
        for key in raw_headers:
            header_map[self._normalize_header(key)] = key

        resolved = {}
        for canonical, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                normalized = self._normalize_header(alias)
                if normalized in header_map:
                    resolved[canonical] = header_map[normalized]
                    break
        return resolved

    def _get_value(self, row, header_map, canonical):
        key = header_map.get(canonical)
        if not key:
            return None
        return row.get(key)

    def _is_empty_row(self, row):
        for value in row.values():
            if str(value).strip():
                return False
        return True

    def _to_float(self, value, row_num, column_name, errors):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.replace(' ', '')
        if ',' in text and '.' in text:
            if text.rfind(',') > text.rfind('.'):
                text = text.replace('.', '').replace(',', '.')
            else:
                text = text.replace(',', '')
        else:
            text = text.replace(',', '.')
        try:
            return float(text)
        except Exception:
            errors.append(_("Fila %s, columna %s: valor invalido '%s'.") % (row_num, column_name, text))
            return None

    def _to_datetime(self, value, row_num, column_name, errors):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        formats = (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%m/%d/%Y",
            "%d/%m/%y",
            "%d-%m-%y",
        )
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt)
            except Exception:
                continue
        try:
            serial = float(text)
            if serial > 20000:
                base = datetime(1899, 12, 30)
                return base + timedelta(days=serial)
        except Exception:
            pass
        errors.append(_("Fila %s, columna %s: fecha invalida '%s'.") % (row_num, column_name, text))
        return None

    def _to_int(self, value, row_num, column_name, errors):
        number = self._to_float(value, row_num, column_name, errors)
        if number is None:
            return None
        if abs(number - round(number)) > 1e-6:
            errors.append(_("Fila %s, columna %s: debe ser entero, valor '%s'.") % (row_num, column_name, value))
            return None
        return int(round(number))

    def _is_positive(self, value):
        return value is not None and value > 0

    def _extract_code(self, value):
        text = (value or '').strip()
        if not text:
            return ''
        match = re.search(r'\[([^\]]+)\]', text)
        if match:
            return match.group(1).strip()
        return text

    def _find_product(self, raw_value, row_num, column_name, errors):
        text = (raw_value or '').strip()
        if not text:
            errors.append(_("Fila %s, columna %s: valor vacio.") % (row_num, column_name))
            return False
        code = self._extract_code(text)
        Product = self.env['product.product']
        product = Product.search([('default_code', '=', code)], limit=2)
        if len(product) == 1:
            return product
        if len(product) > 1:
            errors.append(_("Fila %s, columna %s: codigo '%s' ambiguo.") % (row_num, column_name, code))
            return False
        product = Product.search([('name', 'ilike', text)], limit=2)
        if len(product) == 1:
            return product
        if len(product) > 1:
            errors.append(_("Fila %s, columna %s: nombre '%s' ambiguo.") % (row_num, column_name, text))
            return False
        errors.append(_("Fila %s, columna %s: producto no encontrado '%s'.") % (row_num, column_name, text))
        return False

    def _find_workcenter(self, name, row_num, errors):
        text = (name or '').strip()
        if not text:
            errors.append(_("Fila %s, columna Centro de trabajo: valor vacio.") % row_num)
            return False
        Workcenter = self.env['mrp.workcenter']
        wc = Workcenter.search([('name', '=', text)], limit=2)
        if len(wc) == 1:
            return wc
        if len(wc) > 1:
            errors.append(_("Fila %s: centro de trabajo ambiguo '%s'.") % (row_num, text))
            return False
        wc = Workcenter.search([('name', 'ilike', text)], limit=2)
        if len(wc) == 1:
            return wc
        if len(wc) > 1:
            errors.append(_("Fila %s: centro de trabajo ambiguo '%s'.") % (row_num, text))
            return False
        errors.append(_("Fila %s: centro de trabajo no encontrado '%s'.") % (row_num, text))
        return False

    def _find_partner(self, id_cliente, row_num, errors):
        value = (id_cliente or '').strip()
        if not value:
            errors.append(_("Fila %s: ID CLIENTE vacio.") % row_num)
            return False
        Partner = self.env['res.partner'].with_context(active_test=False)
        if 'vat' in Partner._fields:
            partners = Partner.search([('vat', '=', value)], limit=2)
            if len(partners) == 1:
                return partners
            if len(partners) > 1:
                errors.append(_("Fila %s: ID CLIENTE '%s' es ambiguo (vat).") % (row_num, value))
                return False
        errors.append(_("Fila %s: ID CLIENTE '%s' no encontrado en vat.") % (row_num, value))
        return False

    def action_import_structural(self):
        self.ensure_one()
        lines = self._read_lines()
        if not lines:
            raise UserError(_("El archivo no contiene filas para importar."))

        header_map = self._resolve_headers(lines)
        required_headers = list(self.BASE_REQUIRED_HEADERS)
        if self.create_quotation:
            required_headers.extend(self.QUOTE_REQUIRED_HEADERS)
        missing_headers = [k for k in required_headers if k not in header_map]
        if 'producto' in missing_headers and 'producto_codigo' in header_map:
            missing_headers.remove('producto')
        if missing_headers:
            raise UserError(_("Faltan columnas requeridas en el archivo: %s") % ", ".join(missing_headers))

        errors = []
        groups = []
        current = None

        for idx, row in enumerate(lines):
            row_num = idx + 2
            if self._is_empty_row(row):
                continue

            product_raw = (
                self._get_value(row, header_map, 'producto_codigo')
                or self._get_value(row, header_map, 'producto')
                or ''
            ).strip()
            if product_raw:
                if current:
                    groups.append(current)
                current = {
                    'row_start': row_num,
                    'product_raw': product_raw,
                    'qty_mo': None,
                    'pedido_original': None,
                    'id_cliente': None,
                    'pvp': None,
                    'fecha_cotizacion': None,
                    'largo': None,
                    'ancho': None,
                    'piezas': None,
                    'components': [],
                    'operations': [],
                    'origins': [],
                }
            if not current:
                errors.append(_("Fila %s: no hay PRODUCTO para asociar la fila.") % row_num)
                continue

            qty_mo_raw = (self._get_value(row, header_map, 'cantidad_mo') or '').strip()
            if qty_mo_raw:
                qty_mo = self._to_float(qty_mo_raw, row_num, 'CANTIDAD MO', errors)
                if qty_mo is not None:
                    if qty_mo <= 0:
                        errors.append(_("Fila %s: CANTIDAD MO debe ser mayor que 0.") % row_num)
                    elif current['qty_mo'] is None:
                        current['qty_mo'] = qty_mo
                    elif abs(current['qty_mo'] - qty_mo) > 1e-6:
                        errors.append(_("Fila %s: CANTIDAD MO no coincide con el producto actual.") % row_num)

            pedido_original = (self._get_value(row, header_map, 'pedido_original') or '').strip()
            if pedido_original:
                if current['pedido_original'] is None:
                    current['pedido_original'] = pedido_original
                elif current['pedido_original'] != pedido_original:
                    errors.append(_("Fila %s: Pedido_Original no coincide con el producto actual.") % row_num)

            id_cliente = (self._get_value(row, header_map, 'id_cliente') or '').strip()
            if id_cliente:
                if current['id_cliente'] is None:
                    current['id_cliente'] = id_cliente
                elif current['id_cliente'] != id_cliente:
                    errors.append(_("Fila %s: ID CLIENTE no coincide con el producto actual.") % row_num)

            pvp_raw = (self._get_value(row, header_map, 'pvp') or '').strip()
            if pvp_raw:
                pvp = self._to_float(pvp_raw, row_num, 'PVP', errors)
                if pvp is not None:
                    if pvp < 0:
                        errors.append(_("Fila %s: PVP no puede ser negativo.") % row_num)
                    elif current['pvp'] is None:
                        current['pvp'] = pvp
                    elif abs(current['pvp'] - pvp) > 1e-6:
                        errors.append(_("Fila %s: PVP no coincide con el producto actual.") % row_num)

            fecha_raw = (self._get_value(row, header_map, 'fecha_cotizacion') or '').strip()
            if fecha_raw:
                fecha = self._to_datetime(fecha_raw, row_num, 'FECHA COTIZACION', errors)
                if fecha is not None:
                    if current['fecha_cotizacion'] is None:
                        current['fecha_cotizacion'] = fecha
                    elif current['fecha_cotizacion'] != fecha:
                        errors.append(_("Fila %s: Fecha cotizacion no coincide con el producto actual.") % row_num)

            largo_raw = (self._get_value(row, header_map, 'largo') or '').strip()
            if largo_raw:
                largo_val = self._to_int(largo_raw, row_num, 'LARGO', errors)
                if largo_val is not None:
                    if current['largo'] is None:
                        current['largo'] = largo_val
                    elif current['largo'] != largo_val:
                        errors.append(_("Fila %s: Largo no coincide con el producto actual.") % row_num)

            ancho_raw = (self._get_value(row, header_map, 'ancho') or '').strip()
            if ancho_raw:
                ancho_val = self._to_int(ancho_raw, row_num, 'ANCHO', errors)
                if ancho_val is not None:
                    if current['ancho'] is None:
                        current['ancho'] = ancho_val
                    elif current['ancho'] != ancho_val:
                        errors.append(_("Fila %s: Ancho no coincide con el producto actual.") % row_num)

            piezas_raw = (self._get_value(row, header_map, 'piezas') or '').strip()
            if piezas_raw:
                piezas_val = self._to_int(piezas_raw, row_num, 'PIEZAS', errors)
                if piezas_val is not None:
                    if piezas_val <= 0:
                        errors.append(_("Fila %s: Piezas debe ser mayor que 0.") % row_num)
                    elif current['piezas'] is None:
                        current['piezas'] = piezas_val
                    elif current['piezas'] != piezas_val:
                        errors.append(_("Fila %s: Piezas no coincide con el producto actual.") % row_num)

            comp_raw = (self._get_value(row, header_map, 'componentes_producto') or '').strip()
            if comp_raw:
                comp_qty_raw = (self._get_value(row, header_map, 'por_consumir') or '').strip()
                if not comp_qty_raw:
                    errors.append(_("Fila %s: Por Consumir vacio para componente.") % row_num)
                else:
                    comp_qty = self._to_float(comp_qty_raw, row_num, 'Por Consumir', errors)
                    if comp_qty is not None:
                        if comp_qty <= 0:
                            errors.append(_("Fila %s: Por Consumir debe ser mayor que 0.") % row_num)
                        else:
                            current['components'].append({
                                'raw': comp_raw,
                                'qty': comp_qty,
                                'row_num': row_num,
                            })

            op_raw = (self._get_value(row, header_map, 'operaciones') or '').strip()
            if op_raw:
                wc_raw = (self._get_value(row, header_map, 'centro_trabajo') or '').strip()
                if not wc_raw:
                    errors.append(_("Fila %s: Centro de trabajo vacio para operacion.") % row_num)
                else:
                    current['operations'].append({
                        'name': op_raw,
                        'workcenter': wc_raw,
                        'row_num': row_num,
                    })

            origin_raw = (self._get_value(row, header_map, 'origen') or '').strip()
            if origin_raw and origin_raw not in current['origins']:
                current['origins'].append(origin_raw)

        if current:
            groups.append(current)

        for group in groups:
            row_num = group['row_start']
            largo_val = group.get('largo')
            ancho_val = group.get('ancho')
            piezas_val = group.get('piezas')
            has_any_dim = (
                self._is_positive(largo_val) or
                self._is_positive(ancho_val) or
                self._is_positive(piezas_val)
            )
            has_dims = (
                self._is_positive(largo_val) and
                self._is_positive(ancho_val) and
                self._is_positive(piezas_val)
            )
            area_m2 = None
            if has_dims:
                area_m2 = (largo_val * ancho_val) / 1000000.0
                if area_m2 <= 0:
                    errors.append(_("Fila %s: Largo/Ancho invalidos para calcular m2.") % row_num)
                    group['_skip'] = True
                    continue
                qty_calc = area_m2 * piezas_val
                if group['qty_mo'] is None:
                    group['qty_mo'] = qty_calc
                else:
                    if abs(group['qty_mo'] - qty_calc) > 1e-6:
                        raise UserError(_(
                            "Revisar las cantidades del producto %s ya que no coinciden para la cotizacion y la MO. "
                            "Cantidad Excel: %s, Cantidad calculada: %s (Largo: %s, Ancho: %s, Piezas: %s)."
                        ) % (
                            group.get('product_raw') or row_num,
                            group['qty_mo'],
                            qty_calc,
                            largo_val,
                            ancho_val,
                            piezas_val,
                        ))
                    group['qty_mo'] = qty_calc
            group['_has_any_dim'] = has_any_dim
            group['_has_dims'] = has_dims
            group['_area_m2'] = area_m2

        created_mo_ids = []
        created_bom_ids = []
        created_sale_ids = []
        sale_order_cache = {}

        mrp_production = self.env['mrp.production']
        bom_model = self.env['mrp.bom']
        bom_line_model = self.env['mrp.bom.line']
        routing_model = self.env['mrp.routing.workcenter']
        sale_order_model = self.env['sale.order']
        sale_line_model = self.env['sale.order.line']

        for group in groups:
            row_num = group['row_start']
            if group.get('_skip'):
                continue
            has_any_dim = group.get('_has_any_dim')
            has_dims = group.get('_has_dims')
            area_m2 = group.get('_area_m2')

            if not group['product_raw']:
                errors.append(_("Fila %s: PRODUCTO vacio.") % row_num)
                continue
            if group['qty_mo'] is None:
                errors.append(_("Fila %s: CANTIDAD MO vacia para producto '%s'.") % (row_num, group['product_raw']))
                continue
            if not group['pedido_original']:
                errors.append(_("Fila %s: Pedido_Original vacio para producto '%s'.") % (row_num, group['product_raw']))
                continue
            if not group['components']:
                errors.append(_("Fila %s: sin componentes para producto '%s'.") % (row_num, group['product_raw']))
                continue

            product_col = 'REFERENCIA' if 'producto_codigo' in header_map else 'PRODUCTO'
            product = self._find_product(group['product_raw'], row_num, product_col, errors)
            if not product:
                continue

            component_qty = {}
            for comp in group['components']:
                comp_prod = self._find_product(comp['raw'], comp['row_num'], 'COMPONENTES_PRODUCTO', errors)
                if not comp_prod:
                    continue
                component_qty.setdefault(comp_prod.id, 0.0)
                component_qty[comp_prod.id] += comp['qty']

            if not component_qty:
                errors.append(_("Fila %s: no hay componentes validos para '%s'.") % (row_num, group['product_raw']))
                continue

            bom_line_vals = []
            for prod_id, qty in component_qty.items():
                vals = {
                    'product_id': prod_id,
                    'product_qty': qty,
                }
                if 'product_uom_id' in bom_line_model._fields:
                    vals['product_uom_id'] = self.env['product.product'].browse(prod_id).uom_id.id
                bom_line_vals.append((0, 0, vals))

            bom_vals = {
                'product_tmpl_id': product.product_tmpl_id.id,
                'bom_line_ids': bom_line_vals,
            }
            if 'product_id' in bom_model._fields:
                bom_vals['product_id'] = product.id
            if 'type' in bom_model._fields:
                bom_vals['type'] = 'normal'
            bom = bom_model.create(bom_vals)
            created_bom_ids.append(bom.id)

            seen_ops = set()
            for seq, op in enumerate(group['operations'], start=1):
                op_name = (op.get('name') or '').strip()
                wc_name = (op.get('workcenter') or '').strip()
                if not op_name or not wc_name:
                    continue
                wc = self._find_workcenter(wc_name, op.get('row_num') or row_num, errors)
                if not wc:
                    continue
                key = (op_name.lower(), wc.id)
                if key in seen_ops:
                    continue
                seen_ops.add(key)
                op_vals = {
                    'name': op_name,
                    'workcenter_id': wc.id,
                }
                if 'bom_id' in routing_model._fields:
                    op_vals['bom_id'] = bom.id
                if 'sequence' in routing_model._fields:
                    op_vals['sequence'] = seq
                routing_model.create(op_vals)

            sale_order = False
            sale_line = False
            if self.create_quotation:
                if not group['id_cliente']:
                    errors.append(_("Fila %s: ID CLIENTE requerido para cotizacion.") % row_num)
                    continue
                if group['pvp'] is None:
                    errors.append(_("Fila %s: PVP requerido para cotizacion.") % row_num)
                    continue
                if group['fecha_cotizacion'] is None:
                    errors.append(_("Fila %s: Fecha cotizacion requerida.") % row_num)
                    continue
                if has_any_dim and not has_dims:
                    errors.append(_("Fila %s: Largo/Ancho/Piezas incompletos para cotizacion.") % row_num)
                    continue
                partner = self._find_partner(group['id_cliente'], row_num, errors)
                if not partner:
                    continue
                cache_key = (partner.id, group['pedido_original'])
                sale_order = sale_order_cache.get(cache_key)
                if not sale_order:
                    so_vals = {'partner_id': partner.id}
                    if 'client_order_ref' in sale_order_model._fields:
                        so_vals['client_order_ref'] = group['pedido_original']
                    if 'date_order' in sale_order_model._fields and group['fecha_cotizacion']:
                        so_vals['date_order'] = fields.Datetime.to_string(group['fecha_cotizacion'])
                    sale_order = sale_order_model.create(so_vals)
                    sale_order_cache[cache_key] = sale_order
                    created_sale_ids.append(sale_order.id)

                line_vals = {
                    'order_id': sale_order.id,
                    'product_id': product.id,
                    'product_uom_qty': group['qty_mo'],
                    'price_unit': group['pvp'],
                }
                if 'product_uom' in sale_line_model._fields:
                    line_vals['product_uom'] = product.uom_id.id
                if 'name' in sale_line_model._fields:
                    line_vals['name'] = product.display_name or product.name
                if has_dims and area_m2:
                    line_vals['price_unit'] = group['pvp'] / area_m2
                if group['largo'] is not None and 'x_studio_largo' in sale_line_model._fields:
                    line_vals['x_studio_largo'] = group['largo']
                if group['ancho'] is not None and 'x_studio_ancho' in sale_line_model._fields:
                    line_vals['x_studio_ancho'] = group['ancho']
                if group['piezas'] is not None and 'x_studio_piezas' in sale_line_model._fields:
                    line_vals['x_studio_piezas'] = group['piezas']
                sale_line = sale_line_model.create(line_vals)

            origin_parts = []
            if sale_order:
                origin_parts.append(sale_order.name)
            if group['origins']:
                origin_parts.extend(group['origins'])
            origin = "/".join(origin_parts) if origin_parts else (group['pedido_original'] or '')

            mo_vals = {
                'product_id': product.id,
                'product_qty': group['qty_mo'],
                'bom_id': bom.id,
                'origin': origin,
            }
            if 'x_studio_pedido_original' in mrp_production._fields:
                mo_vals['x_studio_pedido_original'] = group['pedido_original']
            if sale_line and 'sale_line_id' in mrp_production._fields:
                mo_vals['sale_line_id'] = sale_line.id
            mo = mrp_production.create(mo_vals)
            try:
                mo.action_confirm()
            except Exception as exc:
                errors.append(_("Fila %s: error al confirmar MO: %s") % (row_num, str(exc)))
            created_mo_ids.append(mo.id)

        summary = _(
            "Importacion completada. MOs creadas: %s. BOMs creadas: %s. Cotizaciones: %s."
        ) % (
            len(created_mo_ids),
            len(created_bom_ids),
            len(created_sale_ids),
        )
        error_message = "\n".join(errors) if errors else _("Sin errores.")
        result = self.env['mrp.import.result.wizard'].create({
            'summary': summary,
            'error_message': error_message,
            'production_ids': [(6, 0, created_mo_ids)],
            'sale_order_ids': [(6, 0, created_sale_ids)],
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mrp.import.result.wizard',
            'view_mode': 'form',
            'res_id': result.id,
            'target': 'new',
        }
