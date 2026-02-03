# -*- coding: utf-8 -*-
"""Import MOs with Scrap.

This wizard was merged from the standalone module `alterben_mrp_import_mo_with_scrap`.
It creates and confirms Manufacturing Orders (mrp.production) from an Excel (.xls) or CSV file
and optionally records Scrap (stock.scrap) linked to each MO.
"""

from odoo import models, fields, _
from odoo.exceptions import UserError

import base64
import io
import csv

try:
    from odoo.tools.safe_eval import safe_eval
except Exception:  # pragma: no cover
    safe_eval = None

try:
    import xlrd  # For .xls files
except Exception:  # pragma: no cover
    xlrd = None
try:
    import openpyxl  # For .xlsx files
except Exception:  # pragma: no cover
    openpyxl = None


class MrpImportWizard(models.TransientModel):
    _name = 'mrp.import.wizard'
    _description = 'Importador de Órdenes de Fabricación con Desechos'

    file = fields.Binary(string="Archivo Excel/CSV", required=True)
    filename = fields.Char(string="Nombre de archivo")

    def _read_lines(self):
        """Return a list[dict] with the parsed rows."""
        self.ensure_one()
        if not self.file:
            raise UserError(_("Debes adjuntar un archivo para importar."))

        if not self.filename:
            raise UserError(_("Debes indicar el nombre del archivo (filename)."))

        content = base64.b64decode(self.file)
        fname = (self.filename or '').lower()

        if fname.endswith('.csv'):
            csvfile = io.StringIO(content.decode('utf-8-sig'))
            reader = csv.DictReader(csvfile)
            return list(reader)

        if fname.endswith('.xlsx'):
            if openpyxl is None:
                raise UserError(_("No se puede leer Excel .xlsx porque falta la libreria Python 'openpyxl' en el servidor."))
            try:
                workbook = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
            except Exception:
                raise UserError(_("No se pudo leer el archivo .xlsx. Verifica que este bien formado."))
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers_row = next(rows_iter)
            except StopIteration:
                return []
            headers = [str(cell).strip() if cell is not None else '' for cell in headers_row]
            lines = []
            for row in rows_iter:
                line = {}
                for colx in range(len(headers)):
                    value = row[colx] if colx < len(row) else None
                    line[headers[colx]] = str(value).strip() if value is not None else ''
                lines.append(line)
            return lines

        # Excel (.xls)
        if xlrd is None:
            raise UserError(_("No se puede leer Excel .xls porque falta la librería Python 'xlrd' en el servidor."))

        try:
            workbook = xlrd.open_workbook(file_contents=content)
        except Exception:
            raise UserError(_("No se pudo leer el archivo. Verifica que sea .xls o .csv."))
        sheet = workbook.sheet_by_index(0)
        headers = [str(cell.value).strip() for cell in sheet.row(0)]
        lines = []
        for rowx in range(1, sheet.nrows):
            line = {headers[colx]: str(sheet.cell(rowx, colx).value).strip() for colx in range(len(headers))}
            lines.append(line)
        return lines

    def _get_header_map(self, lines):
        header_map = {}
        if not lines:
            return header_map
        for key in lines[0].keys():
            if isinstance(key, str):
                header_map[key.strip().lower()] = key
        return header_map

    def _resolve_xml_id(self, xml_id, expected_model, row_num, column_name, errors):
        xml_id = (xml_id or '').strip()
        if not xml_id:
            errors.append(_("Fila %s, columna %s: valor vacío.") % (row_num, column_name))
            return False

        if '.' in xml_id:
            try:
                record = self.env.ref(xml_id)
            except ValueError:
                errors.append(_("Fila %s, columna %s: xml_id '%s' no existe.") % (row_num, column_name, xml_id))
                return False
        else:
            data_recs = self.env['ir.model.data'].search([
                ('name', '=', xml_id),
                ('model', '=', expected_model),
            ])
            if not data_recs:
                errors.append(_("Fila %s, columna %s: xml_id '%s' no existe.") % (row_num, column_name, xml_id))
                return False
            if len(data_recs) > 1:
                modules = ', '.join(sorted(set(data_recs.mapped('module'))))
                errors.append(_(
                    "Fila %s, columna %s: xml_id '%s' es ambiguo (módulos: %s)."
                ) % (row_num, column_name, xml_id, modules))
                return False
            record = self.env[expected_model].browse(data_recs.res_id)

        if not record or not record.exists():
            errors.append(_("Fila %s, columna %s: xml_id '%s' no existe.") % (row_num, column_name, xml_id))
            return False
        if record._name != expected_model:
            errors.append(_(
                "Fila %s, columna %s: xml_id '%s' no corresponde a %s."
            ) % (row_num, column_name, xml_id, expected_model))
            return False
        return record

    def _get_import_master_type(self):
        """Return the configured Master Type for imports (if any)."""
        self.ensure_one()
        mtype = self.env['mrp.master.type'].search([
            ('active', '=', True),
            ('import_scrap_location_id', '!=', False),
        ], limit=1)
        if not mtype:
            mtype = self.env['mrp.master.type'].search([
                ('active', '=', True),
            ], limit=1)
        return mtype

    def _get_scrap_location(self):
        """Return the configured scrap location for imports."""
        self.ensure_one()
        mtype = self._get_import_master_type()
        if mtype and mtype.import_scrap_location_id:
            return mtype.import_scrap_location_id
        raise UserError(_("Configure Ubicación de Desecho (Importación OF) en Parámetros de Órdenes Maestras."))

    def _get_scrap_origin_field(self):
        fields_map = self.env['stock.scrap']._fields
        if 'origin' in fields_map:
            return 'origin'
        if 'reference' in fields_map:
            return 'reference'
        if 'name' in fields_map:
            return 'name'
        return False

    def _get_scrap_source_location(self, mo):
        if 'location_src_id' in mo._fields and mo.location_src_id:
            return mo.location_src_id
        if 'location_id' in mo._fields and mo.location_id:
            return mo.location_id
        if 'picking_type_id' in mo._fields and mo.picking_type_id and mo.picking_type_id.default_location_src_id:
            return mo.picking_type_id.default_location_src_id

        company = mo.company_id if 'company_id' in mo._fields and mo.company_id else self.env.company
        warehouse = self.env['stock.warehouse'].search([
            ('company_id', '=', company.id),
        ], limit=1)
        if warehouse and warehouse.lot_stock_id:
            return warehouse.lot_stock_id

        location = self.env['stock.location'].search([
            '|', ('company_id', '=', company.id), ('company_id', '=', False),
            ('usage', '=', 'internal'),
        ], order='id', limit=1)
        return location
    def action_import_mrp(self):
        self.ensure_one()
        lines = self._read_lines()

        errores = []
        mrp_production = self.env['mrp.production']
        stock_scrap = self.env['stock.scrap']
        scrap_location = self._get_scrap_location()
        import_conf = self._get_import_master_type()
        auto_validate_import_scrap = bool(import_conf and import_conf.auto_validate_import_scrap)
        if not auto_validate_import_scrap:
            auto_validate_import_scrap = bool(self.env['mrp.master.type'].search([
                ('auto_validate_import_scrap', '=', True),
            ], limit=1))
        allow_import_scrap_without_stock = bool(import_conf and import_conf.allow_import_scrap_without_stock)
        if not allow_import_scrap_without_stock:
            allow_import_scrap_without_stock = bool(self.env['mrp.master.type'].search([
                ('allow_import_scrap_without_stock', '=', True),
            ], limit=1))
        scrap_origin_field = self._get_scrap_origin_field()
        created_mo_ids = []
        created_scrap_ids = []
        header_map = self._get_header_map(lines)

        if not lines:
            raise UserError(_("El archivo no contiene filas para importar."))

        required_headers = [
            'PRODUCTO/ID EXTERNO',
            'PRODUCT_QTY',
            'BOM_ID/ID',
            'ORIGEN',
            'PEDIDO_ORIGINAL',
            'SCRAP_QTY',
        ]
        missing_headers = [h for h in required_headers if h.lower() not in header_map]
        if missing_headers:
            raise UserError(_(
                "Faltan columnas requeridas en el archivo: %s"
            ) % ", ".join(missing_headers))

        def _get_value(row, header):
            key = header_map.get(header.lower())
            if not key:
                return None
            return row.get(key)

        def _set_auto_validate_info(scrap_rec, failed, message=None):
            vals = {}
            if 'auto_validate_failed' in stock_scrap._fields:
                vals['auto_validate_failed'] = failed
            if 'auto_validate_message' in stock_scrap._fields:
                vals['auto_validate_message'] = message or False
            if vals:
                scrap_rec.sudo().write(vals)

        def _get_scrap_qty(scrap_rec):
            for fname in ('scrap_qty', 'quantity', 'product_qty'):
                if fname in stock_scrap._fields:
                    return getattr(scrap_rec, fname) or 0.0
            return 0.0

        def _get_available_qty(scrap_rec):
            quant = self.env['stock.quant']
            if not hasattr(quant, '_get_available_quantity'):
                return None
            if not scrap_rec.product_id or not scrap_rec.location_id:
                return None
            lot = scrap_rec.lot_id if 'lot_id' in stock_scrap._fields else False
            owner = scrap_rec.owner_id if 'owner_id' in stock_scrap._fields else False
            package = scrap_rec.package_id if 'package_id' in stock_scrap._fields else False
            try:
                return quant._get_available_quantity(
                    scrap_rec.product_id,
                    scrap_rec.location_id,
                    lot_id=lot,
                    owner_id=owner,
                    package_id=package,
                )
            except Exception:
                return None

        def _build_auto_validate_reason(scrap_rec):
            reasons = []
            if 'location_id' in stock_scrap._fields and not scrap_rec.location_id:
                reasons.append(_("Sin ubicacion de origen."))
            if 'scrap_location_id' in stock_scrap._fields and not scrap_rec.scrap_location_id:
                reasons.append(_("Sin ubicacion de desecho."))
            qty = _get_scrap_qty(scrap_rec)
            if qty <= 0:
                reasons.append(_("Cantidad de desecho <= 0."))
            if 'product_uom_id' in stock_scrap._fields and not scrap_rec.product_uom_id:
                reasons.append(_("Sin unidad de medida."))
            if 'uom_id' in stock_scrap._fields and not scrap_rec.uom_id:
                reasons.append(_("Sin unidad de medida."))
            if scrap_rec.product_id and scrap_rec.product_id.tracking != 'none':
                lot_id = scrap_rec.lot_id if 'lot_id' in stock_scrap._fields else False
                lot_name = scrap_rec.lot_name if 'lot_name' in stock_scrap._fields else False
                if not lot_id and not lot_name:
                    reasons.append(_("Falta lote/serie."))
            if 'move_id' in stock_scrap._fields:
                if scrap_rec.move_id:
                    move_label = scrap_rec.move_id.name or str(scrap_rec.move_id.id)
                    reasons.append(_("Movimiento %s estado %s.") % (move_label, scrap_rec.move_id.state))
                else:
                    reasons.append(_("Sin movimiento de desecho."))
            available_qty = _get_available_qty(scrap_rec)
            if available_qty is not None:
                reasons.append(_("Disponible en origen: %s.") % available_qty)
            return " ".join(reasons) if reasons else _("Estado en borrador despues de validar.")

        def _force_confirm_action(scrap_rec, action):
            if not isinstance(action, dict):
                return action
            res_model = action.get('res_model')
            if not res_model:
                res_model = 'stock.scrap.confirmation'
            ctx = action.get('context') or {}
            if isinstance(ctx, str) and safe_eval:
                try:
                    ctx = safe_eval(ctx, {'uid': self.env.uid})
                except Exception:
                    ctx = {}
            if not isinstance(ctx, dict):
                ctx = {}
            ctx.setdefault('active_id', scrap_rec.id)
            ctx.setdefault('active_ids', [scrap_rec.id])
            ctx.setdefault('active_model', 'stock.scrap')
            ctx.setdefault('default_scrap_id', scrap_rec.id)
            try:
                model = self.env[res_model]
            except Exception:
                model = None
            if not model:
                return action
            if action.get('res_id'):
                wiz = model.with_context(ctx).browse(action['res_id'])
            else:
                wiz = model.with_context(ctx).create({})
            for method in ('action_confirm', 'confirm', 'action_validate', 'button_confirm', 'process'):
                if hasattr(wiz, method):
                    return getattr(wiz, method)()
            if res_model != 'stock.scrap.confirmation':
                fallback = {
                    'res_model': 'stock.scrap.confirmation',
                    'context': ctx,
                }
                return _force_confirm_action(scrap_rec, fallback)
            return action

        report_auto_validate_in_summary = False
        for idx, row in enumerate(lines):
            try:
                row_num = idx + 2
                if not any((str(value).strip() if value is not None else '') for value in row.values()):
                    continue

                product_xmlid = (_get_value(row, 'PRODUCTO/ID EXTERNO') or '').strip()
                bom_xmlid = (_get_value(row, 'BOM_ID/ID') or '').strip()
                origin = (_get_value(row, 'ORIGEN') or '').strip()
                pedido_original = (_get_value(row, 'PEDIDO_ORIGINAL') or '').strip()
                qty_raw = (_get_value(row, 'PRODUCT_QTY') or '').strip()
                scrap_qty_raw = (_get_value(row, 'SCRAP_QTY') or '').strip()

                if not product_xmlid:
                    errores.append(_("Fila %s, columna PRODUCTO/ID EXTERNO: valor vacío.") % row_num)
                    continue
                if not bom_xmlid:
                    errores.append(_("Fila %s, columna BOM_ID/ID: valor vacío.") % row_num)
                    continue
                if not origin:
                    errores.append(_("Fila %s, columna ORIGEN: valor vacío.") % row_num)
                    continue
                if not pedido_original:
                    errores.append(_("Fila %s, columna PEDIDO_ORIGINAL: valor vacío.") % row_num)
                    continue
                if not qty_raw:
                    errores.append(_("Fila %s, columna PRODUCT_QTY: valor vacío.") % row_num)
                    continue

                try:
                    qty = float(qty_raw)
                except Exception:
                    errores.append(_("Fila %s, columna PRODUCT_QTY: valor inválido '%s'.") % (row_num, qty_raw))
                    continue
                if qty <= 0:
                    errores.append(_("Fila %s, columna PRODUCT_QTY: debe ser mayor que 0.") % row_num)
                    continue

                scrap_qty = 0.0
                if scrap_qty_raw:
                    try:
                        scrap_qty = float(scrap_qty_raw)
                    except Exception:
                        errores.append(_("Fila %s, columna SCRAP_QTY: valor inválido '%s'.") % (row_num, scrap_qty_raw))
                        continue
                    if scrap_qty < 0:
                        errores.append(_("Fila %s, columna SCRAP_QTY: no puede ser negativo.") % row_num)
                        continue

                product = self._resolve_xml_id(product_xmlid, 'product.product', row_num, 'PRODUCTO/ID EXTERNO', errores)
                if not product:
                    continue

                bom = self._resolve_xml_id(bom_xmlid, 'mrp.bom', row_num, 'BOM_ID/ID', errores)
                if not bom:
                    continue

                if 'x_studio_pedido_original' not in mrp_production._fields:
                    errores.append(_("Fila %s: Campo x_studio_pedido_original no existe en mrp.production.") % row_num)
                    continue

                bom_lines = bom.bom_line_ids.sorted(key=lambda line: (line.sequence, line.id))
                if not bom_lines:
                    errores.append(_(
                        "Fila %s: La BoM %s no tiene componentes para derivar producto de desecho."
                    ) % (row_num, bom_xmlid))
                    continue

                scrap_product = bom_lines[0].product_id
                if not scrap_product:
                    errores.append(_(
                        "Fila %s: La BoM %s no tiene producto de componente válido."
                    ) % (row_num, bom_xmlid))
                    continue

                if scrap_qty > 0 and 'x_studio_pedido_original' not in stock_scrap._fields:
                    errores.append(_("Fila %s: Campo x_studio_pedido_original no existe en stock.scrap.") % row_num)
                    continue

                if scrap_qty > 0 and not scrap_origin_field:
                    errores.append(_("Fila %s: No existe campo origin/reference/name en stock.scrap.") % row_num)
                    continue

                mo_vals = {
                    'product_id': product.id,
                    'product_qty': qty,
                    'bom_id': bom.id,
                    'origin': origin,
                    'x_studio_pedido_original': pedido_original,
                }
                mo = mrp_production.create(mo_vals)
                mo.action_confirm()
                created_mo_ids.append(mo.id)

                if scrap_qty > 0:
                    source_location = self._get_scrap_source_location(mo)
                    if not source_location:
                        errores.append(_("Fila %s: No se encontró ubicación de origen para el desecho.") % row_num)
                        mo.unlink()
                        continue

                    scrap_vals = {
                        'product_id': scrap_product.id,
                        'production_id': mo.id,
                        'location_id': source_location.id,
                        'scrap_location_id': scrap_location.id,
                    }
                    if 'scrap_qty' in stock_scrap._fields:
                        scrap_vals['scrap_qty'] = scrap_qty
                    if 'quantity' in stock_scrap._fields:
                        scrap_vals['quantity'] = scrap_qty
                    if 'product_uom_id' in stock_scrap._fields:
                        scrap_vals['product_uom_id'] = scrap_product.uom_id.id
                    elif 'uom_id' in stock_scrap._fields:
                        scrap_vals['uom_id'] = scrap_product.uom_id.id
                    if 'company_id' in stock_scrap._fields and getattr(mo, 'company_id', False):
                        scrap_vals['company_id'] = mo.company_id.id
                    if scrap_origin_field:
                        scrap_vals[scrap_origin_field] = origin
                    for field_name in stock_scrap._fields:
                        if field_name.startswith('x_studio_origen'):
                            scrap_vals[field_name] = origin
                    if 'x_studio_pedido_original' in stock_scrap._fields:
                        scrap_vals['x_studio_pedido_original'] = pedido_original

                    scrap = stock_scrap.create(scrap_vals)
                    created_scrap_ids.append(scrap.id)
                    scrap_qty_value = _get_scrap_qty(scrap)
                    available_qty = _get_available_qty(scrap)
                    if available_qty is None:
                        stock_status = _("Stock desconocido.")
                    elif available_qty >= scrap_qty_value:
                        stock_status = _("Stock suficiente.")
                    else:
                        stock_status = _("Stock insuficiente (disp: %s).") % available_qty
                    if 'import_stock_status' in stock_scrap._fields:
                        scrap.sudo().write({'import_stock_status': stock_status})
                    insufficient_stock = available_qty is not None and scrap_qty_value > available_qty
                    if auto_validate_import_scrap and scrap.state == 'draft':
                        def _handle_validation_error(scrap_rec, exception_obj, row_num_str):
                            """Build a detailed error message for scrap validation failure."""
                            available_qty = _get_available_qty(scrap_rec)
                            scrap_qty = _get_scrap_qty(scrap_rec)
                            error_details = []
                            msg = getattr(exception_obj, 'name', str(exception_obj))
                            error_details.append(msg)
                            if available_qty is not None:
                                location_name = scrap_rec.location_id.display_name if scrap_rec.location_id else _("Ubicación de origen desconocida")
                                error_details.append(
                                    _("Cantidad a desechar: %s. Cantidad disponible en '%s': %s.") %
                                    (scrap_qty, location_name, available_qty)
                                )
                            if available_qty is not None and scrap_qty > available_qty:
                                error_details.append(_("CAUSA PROBABLE: No hay suficiente stock disponible."))
                            final_reason = " ".join(error_details)
                            if report_auto_validate_in_summary:
                                errores.append(_("Fila %s: Error al validar desecho: %s") % (row_num_str, final_reason))
                            _set_auto_validate_info(scrap_rec, True, final_reason)

                        auto_validate_reason = False
                        try:
                            requires_lot = scrap.product_id.tracking != 'none'
                            lot_id = scrap.lot_id if 'lot_id' in stock_scrap._fields else False
                            lot_name = scrap.lot_name if 'lot_name' in stock_scrap._fields else False
                            if requires_lot and not lot_id and not lot_name:
                                auto_validate_reason = _("Falta lote/serie para validar automaticamente.")
                                if report_auto_validate_in_summary:
                                    errores.append(_("Fila %s: %s") % (row_num, auto_validate_reason))
                                _set_auto_validate_info(scrap, True, auto_validate_reason)
                                continue
                            if insufficient_stock and not allow_import_scrap_without_stock:
                                auto_validate_reason = _("Stock insuficiente; validar sin stock desactivado.")
                                if report_auto_validate_in_summary:
                                    errores.append(_("Fila %s: %s") % (row_num, auto_validate_reason))
                                _set_auto_validate_info(scrap, True, auto_validate_reason)
                                continue
                            if hasattr(scrap, 'action_validate'):
                                scrap_ctx = dict(self.env.context, not_unlink_on_discard=True)
                                if allow_import_scrap_without_stock:
                                    scrap_ctx.update({
                                        'skip_warning': True,
                                        'skip_stock_warning': True,
                                        'scrap_qty_over': True,
                                    })
                                result = scrap.sudo().with_context(scrap_ctx).action_validate()
                                if isinstance(result, dict) and allow_import_scrap_without_stock:
                                    _force_confirm_action(scrap, result)
                            else:
                                auto_validate_reason = _("El modelo stock.scrap no tiene metodo action_validate.")
                                if report_auto_validate_in_summary:
                                    errores.append(_("Fila %s: %s") % (row_num, auto_validate_reason))
                                _set_auto_validate_info(scrap, True, auto_validate_reason)
                                continue
                            refreshed = stock_scrap.browse(scrap.id)
                            if refreshed.state == 'draft' and hasattr(refreshed, '_action_validate'):
                                try:
                                    refreshed.sudo()._action_validate()
                                except Exception as e:
                                    _handle_validation_error(scrap, e, row_num)
                                    continue
                            refreshed = stock_scrap.browse(scrap.id)
                            if refreshed.state == 'draft' and hasattr(refreshed, '_create_scrap_move'):
                                move_exists = bool(getattr(refreshed, 'move_id', False)) if 'move_id' in stock_scrap._fields else False
                                if not move_exists:
                                    try:
                                        refreshed.sudo()._create_scrap_move()
                                        refreshed.sudo().write({'state': 'done'})
                                    except Exception as e:
                                        _handle_validation_error(scrap, e, row_num)
                                        continue
                        except Exception as e:
                            _handle_validation_error(scrap, e, row_num)
                            continue
                        refreshed = stock_scrap.browse(scrap.id)
                        if refreshed.state == 'draft':
                            auto_validate_reason = _build_auto_validate_reason(refreshed)
                            if report_auto_validate_in_summary:
                                errores.append(_("Fila %s: El desecho no se validó automáticamente: %s") % (row_num, auto_validate_reason))
                            _set_auto_validate_info(scrap, True, auto_validate_reason)
                        else:
                            _set_auto_validate_info(scrap, False, False)
                    if False:
                        try:
                            if scrap.product_id.tracking != 'none' and not scrap.lot_id:
                                errores.append(_(
                                    "Fila %s: El desecho requiere lote/serie para validarse automáticamente."
                                ) % row_num)
                            else:
                                if hasattr(scrap, 'action_validate'):
                                    scrap.sudo().action_validate()
                                else:
                                    errores.append(_(
                                        "Fila %s: El modelo stock.scrap no tiene metodo action_validate."
                                    ) % row_num)
                                    continue
                                refreshed = stock_scrap.browse(scrap.id)
                                if refreshed.state == 'draft' and hasattr(refreshed, '_action_validate'):
                                    try:
                                        refreshed.sudo()._action_validate()
                                    except Exception as e:
                                        errores.append(_("Fila %s: Error al validar desecho: %s") % (row_num, str(e)))
                                        continue
                        except Exception as e:
                            errores.append(_("Fila %s: Error al validar desecho: %s") % (row_num, str(e)))
                            continue
                        refreshed = stock_scrap.browse(scrap.id)
                        if refreshed.state == 'draft':
                            errores.append(_("Fila %s: El desecho no se validó automáticamente.") % row_num)
            except Exception as e:
                errores.append(_("Fila %s: %s") % (idx + 2, str(e)))

        summary = _("Importación completada. MOs creadas: %s. Desechos creados: %s. Auto-validación: %s.") % (
            len(created_mo_ids),
            len(created_scrap_ids),
            _("Sí") if auto_validate_import_scrap else _("No"),
        )
        error_message = "\n".join(errores) if errores else _("Sin errores.")
        result = self.env['mrp.import.result.wizard'].create({
            'summary': summary,
            'error_message': error_message,
            'production_ids': [(6, 0, created_mo_ids)],
            'scrap_ids': [(6, 0, created_scrap_ids)],
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mrp.import.result.wizard',
            'view_mode': 'form',
            'res_id': result.id,
            'target': 'new',
        }
